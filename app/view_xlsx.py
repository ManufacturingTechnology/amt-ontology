"""
app.view_xlsx
=============
Export a view ontology (currently ``visitor_view`` — the "Product Interest
Category" taxonomy) as a multi-sheet xlsx mirroring the historical
``resources/AMT Taxonomy - Product Interest Category.xlsx`` workbook.

This is an *additive* export — it does not replace
:mod:`app.view_csv`, which continues to produce the flat three-column CSV
that the browser's Download button serves and that downstream systems
(Maritz, Hubspot, the data warehouse) ingest. The xlsx is for human-facing
governance: it reproduces the spreadsheet AMT's taxonomy stewards have used
historically, but driven by the ontology rather than maintained by hand.

Sheets generated
----------------

1. **Taxonomy Metadata** — name, owner, status, governors, consumers, tech
   stewards, related products/systems, next review date, related
   documentation, use cases, notes. Sourced from the view ontology's
   ``dcterms:title`` and ``amtmeta:*`` annotation properties on its
   ``owl:Ontology`` header.

2. **Release Log** — release number, date, scope summary, owner, approval
   status, release notes URL. Sourced from ``owl:versionInfo`` plus the
   ``amtmeta:release*`` annotation properties. Terms-added/-modified/
   -deprecated columns are intentionally left blank — those are derivable
   from ``git diff`` against the previous tag and belong in the linked
   GitHub release body.

3. **Hierarchy & Mapping** — Tier 1 / Tier 2 / Tier 3 hierarchy mirroring
   :func:`app.view_csv.generate_view_csv_rows` output, with the Excel-
   compatible "Current version" marker on row 1 and the historical Entry
   ID / Notes columns left empty (the Entry ID concept is subsumed by
   stable IRIs in the ontology and is no longer maintained).

4. **Look-up values** — Approval Status options sourced from each
   ``amtmeta:Status`` individual's ``skos:prefLabel`` (Draft, Pending
   Review, Approved, Deprecated). Emitted in the canonical Excel order.

The previously-versioned ``Previous versions`` and ``Hierarchy & Mapping
(SAMPLE)`` sheets are *not* generated — old versions are recoverable from
``git checkout <tag>`` followed by ``make dist-xlsx``, so a separate
sheet for them in the current workbook is redundant.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from urllib.parse import unquote, urlparse

import rdflib
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from rdflib.namespace import DCTERMS, OWL, RDF, RDFS, SKOS

from .namespaces import NS_AMTMETA
from .view_csv import generate_view_csv_rows

# ---------------------------------------------------------------------------
# Hyperlink styling
# ---------------------------------------------------------------------------

#: Excel's stock hyperlink font — Office blue (0563C1) with single underline.
#: Applied by :func:`_set_hyperlink` so URL cells render as clickable
#: hyperlinks matching the look of the historical resource workbook.
HYPERLINK_FONT = Font(color="0563C1", underline="single")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

#: Excel canonical ordering for the Look-up values sheet.
APPROVAL_STATUS_ORDER: tuple[str, ...] = ("Draft", "Pending Review", "Approved", "Deprecated")

#: Filename for the dist export — kept identical to the historical Excel name
#: so downstream consumers that fetch the artefact by filename keep working.
DIST_FILENAME = "AMT Taxonomy - Product Interest Category.xlsx"

#: Headers for the Release Log sheet, in the same order as the resource workbook.
RELEASE_LOG_HEADERS: tuple[str, ...] = (
    "Release Number",
    "Release Date",
    "Scope Summary",
    "Total Terms Added",
    "Terms Modified",
    "Terms Deprecated",
    "Owner",
    "Approval Status",
    "Release notes",
)


# ---------------------------------------------------------------------------
# Graph-walking helpers
# ---------------------------------------------------------------------------


def _find_ontology_iri(graph: rdflib.Graph) -> rdflib.URIRef | None:
    """Return the first ``owl:Ontology`` subject in *graph*.

    Each view file declares exactly one such subject (the view's IRI), so
    this is well-defined for our inputs. Returns ``None`` for the degenerate
    case of an empty / malformed graph.
    """
    for s in graph.subjects(RDF.type, OWL.Ontology):
        if isinstance(s, rdflib.URIRef):
            return s
    return None


def _collect_values(graph: rdflib.Graph, subject, prop) -> list[str]:
    """All literal/IRI objects of ``(subject, prop, ?)``, stringified.

    Order is rdflib's iteration order — not guaranteed stable across runs,
    which is fine for set-comparison-based tests; if a stable order is
    needed callers can sort the result themselves.
    """
    return [str(o) for o in graph.objects(subject, prop)]


def _first_value(graph: rdflib.Graph, subject, prop, default: str = "") -> str:
    """First object of ``(subject, prop, ?)`` as a string, or *default*."""
    for o in graph.objects(subject, prop):
        return str(o)
    return default


def _resolve_status_label(amtmeta_graph: rdflib.Graph, status) -> str:
    """Render a ``amtmeta:Status`` IRI as its ``rdfs:label`` (e.g. "Active").

    Literals are returned verbatim; IRIs fall back to the local name if no
    label is asserted in *amtmeta_graph*.
    """
    if not isinstance(status, rdflib.URIRef):
        return str(status)
    for lit in amtmeta_graph.objects(status, RDFS.label):
        return str(lit)
    return str(status).rsplit("#", 1)[-1]


# ---------------------------------------------------------------------------
# Sheet populators
# ---------------------------------------------------------------------------


def _bold_first_cell(ws, row_index: int) -> None:
    """Apply bold to col 1 of *row_index* — used for field labels."""
    ws.cell(row_index, 1).font = Font(bold=True)


def _populate_metadata_sheet(
    ws,
    view_graph: rdflib.Graph,
    amtmeta_graph: rdflib.Graph,
) -> None:
    """Taxonomy Metadata sheet — field name in col 1, value(s) in col 2.

    Use cases and Notes can span multiple rows (one row per item); the
    first row carries the label, continuation rows leave col 1 blank. This
    matches the resource workbook's manual editing convention.
    """
    ont = _find_ontology_iri(view_graph)
    if ont is None:
        return

    title = _first_value(view_graph, ont, DCTERMS.title)
    status_iri = next(view_graph.objects(ont, NS_AMTMETA.status), None)
    status_label = _resolve_status_label(amtmeta_graph, status_iri) if status_iri else ""

    # Single-value rows — joined with ", " when the property has multiple
    # asserted values (e.g. multiple governors).
    single_value_rows: list[tuple[str, str]] = [
        ("Taxonomy name:", title),
        ("Taxonomy owner:", ", ".join(_collect_values(view_graph, ont, NS_AMTMETA.taxonomyOwner))),
        ("Status:", status_label),
        ("Governers:", ", ".join(_collect_values(view_graph, ont, NS_AMTMETA.governedBy))),
        ("Consumers:", ", ".join(_collect_values(view_graph, ont, NS_AMTMETA.consumedBy))),
        (
            "Tech stewards:",
            ", ".join(_collect_values(view_graph, ont, NS_AMTMETA.technicalSteward)),
        ),
        (
            "Related products:",
            ", ".join(_collect_values(view_graph, ont, NS_AMTMETA.relatedProduct)),
        ),
        (
            "Related systems/databases:",
            ", ".join(_collect_values(view_graph, ont, NS_AMTMETA.relatedSystem)),
        ),
        ("Next update date:", _first_value(view_graph, ont, NS_AMTMETA.nextReviewDate)),
        ("Related documentation:", _first_value(view_graph, ont, NS_AMTMETA.relatedDocumentation)),
    ]

    for label, value in single_value_rows:
        ws.append([label, value])
        _bold_first_cell(ws, ws.max_row)

    # Multi-row groups — first row carries the label, continuation rows
    # leave col 1 blank.
    _append_multi_row_field(ws, "Use cases:", _collect_values(view_graph, ont, NS_AMTMETA.useCase))
    _append_multi_row_field(ws, "Notes:", _collect_values(view_graph, ont, NS_AMTMETA.notes))


def _append_multi_row_field(ws, label: str, values: list[str]) -> None:
    """Emit a field whose value spans one row per element of *values*.

    When *values* is empty a single row is emitted with an empty value
    cell, so the field label is still present in the sheet.
    """
    if not values:
        ws.append([label, ""])
        _bold_first_cell(ws, ws.max_row)
        return
    ws.append([label, values[0]])
    _bold_first_cell(ws, ws.max_row)
    for v in values[1:]:
        ws.append(["", v])


def _populate_release_log_sheet(
    ws,
    view_graph: rdflib.Graph,
    amtmeta_graph: rdflib.Graph,
) -> None:
    """Release Log sheet — one header row, one data row for the current release.

    Terms-added / -modified / -deprecated columns are left blank: those
    derive from ``git diff`` against the previous tag and live in the
    GitHub release body, not in static ontology metadata (see amtmeta.ttl
    D5).
    """
    ws.append(list(RELEASE_LOG_HEADERS))
    for cell in ws[1]:
        cell.font = Font(bold=True)

    ont = _find_ontology_iri(view_graph)
    if ont is None:
        return

    release_number = _first_value(view_graph, ont, OWL.versionInfo)
    release_date = _first_value(view_graph, ont, NS_AMTMETA.releaseDate)
    scope = _first_value(view_graph, ont, NS_AMTMETA.scopeSummary)
    owner = _first_value(view_graph, ont, NS_AMTMETA.releaseOwner)
    approval_iri = next(view_graph.objects(ont, NS_AMTMETA.approvalStatus), None)
    approval = _resolve_status_label(amtmeta_graph, approval_iri) if approval_iri else ""
    notes = _first_value(view_graph, ont, NS_AMTMETA.releaseNotes)

    ws.append(
        [
            release_number,
            release_date,
            scope,
            None,  # Total Terms Added — derive from git
            None,  # Terms Modified — derive from git
            None,  # Terms Deprecated — derive from git
            owner,
            approval,
            notes,
        ]
    )


def _populate_hierarchy_sheet(
    ws,
    view_graph: rdflib.Graph,
    source_graph: rdflib.Graph,
) -> None:
    """Hierarchy & Mapping sheet — Tier 1 / Tier 2 / Tier 3 rows.

    Row 1 carries the "Current version" marker (matches the resource
    workbook). Row 2 is the historical header (``Entry ID, Tier 1, Tier 2,
    Tier 3, Notes``). Subsequent rows hold the data, with Entry ID and
    Notes left blank — Entry ID's role is filled by stable ontology IRIs
    instead, and Notes are not collected per-leaf.
    """
    ws.append(["Current version"])
    ws.append(["Entry ID", "Tier 1", "Tier 2", "Tier 3", "Notes"])
    for cell in ws[2]:
        cell.font = Font(bold=True)

    _, rows = generate_view_csv_rows(view_graph, source_graph)
    for major, sub, product in rows:
        ws.append([None, major, sub, product, None])


def _populate_lookup_sheet(ws, amtmeta_graph: rdflib.Graph) -> None:
    """Look-up values sheet — Approval Status options.

    Each ``amtmeta:Status`` individual carries a ``skos:prefLabel`` that
    matches one of the canonical Excel values (Draft / Pending Review /
    Approved / Deprecated). They're emitted in the canonical Excel order
    (:data:`APPROVAL_STATUS_ORDER`) so the generated sheet aligns visually
    with the historical workbook; any extras land at the end in
    alphabetical order.
    """
    ws.append(["Approval Status Options"])
    ws.cell(1, 1).font = Font(bold=True)

    pref_labels: list[str] = []
    for status in amtmeta_graph.subjects(RDF.type, NS_AMTMETA.Status):
        for lit in amtmeta_graph.objects(status, SKOS.prefLabel):
            pref_labels.append(str(lit))
            break

    ordered = [v for v in APPROVAL_STATUS_ORDER if v in pref_labels]
    extras = sorted(set(pref_labels) - set(ordered))
    for value in ordered + extras:
        ws.append([value])


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------


def _autosize_columns(ws, *, min_width: int = 8, max_width: int = 120) -> None:
    """Size each column so the widest cell content fits.

    Bounds the resulting width between *min_width* and *max_width* so a
    single long value (e.g. the SharePoint URL or a paragraph-long use
    case) doesn't blow the column out to absurd widths. For multi-line
    cells, only the longest line is considered (Excel renders wrapped
    text vertically anyway).

    Deterministic — width is derived from cell content, which is itself
    sourced from the ontology and stable across runs. Safe to call after
    populating but before any post-save zip rewrite.
    """
    for col in ws.iter_cols():
        if not col:
            continue
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            text = str(cell.value)
            # Multi-line cells: pick the longest individual line.
            for line in text.splitlines() or [text]:
                if len(line) > max_len:
                    max_len = len(line)
        width = min(max(min_width, max_len + 2), max_width)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


def _wrap_column(ws, column: int) -> None:
    """Apply ``wrap_text=True`` to every cell in *column* of *ws*.

    Used on free-form columns (the Taxonomy Metadata value column, the
    Release Log scope-summary column) where contents can run to multiple
    sentences or include long URLs. Combined with :func:`_autosize_columns`
    capping at ``max_width``, this keeps the workbook readable in Excel
    without forcing the user to widen columns by hand.
    """
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_col=column, max_col=column):
        for cell in row:
            cell.alignment = wrap


def _set_hyperlink(cell, url: str) -> None:
    """Render *cell* as an Excel hyperlink to *url*.

    Display text is the URL's filename segment (the last path component,
    URL-decoded so ``%20`` becomes a space). For a SharePoint URL like
    ``https://…/Documents/Taxonomy%20Change%20Process.docx?…``, the cell
    shows ``Taxonomy Change Process.docx`` and clicking it opens the
    underlying URL — same UX as the historical resource workbook's
    hyperlinked filename. Applies the stock Office hyperlink font
    (blue + underline).
    """
    parsed = urlparse(url)
    filename = unquote(parsed.path.rsplit("/", 1)[-1])
    cell.value = filename or url
    cell.hyperlink = url
    cell.font = HYPERLINK_FONT


def _hyperlinkify_url_cells(ws) -> None:
    """Walk every populated cell; convert any whose string value is an
    ``http://`` / ``https://`` URL into a clickable hyperlink via
    :func:`_set_hyperlink`.

    Applied per-sheet after population so cells like Taxonomy Metadata's
    ``Related documentation:`` row and Release Log's ``Release notes`` and
    ``GitHub release`` URLs render with a friendly filename / tag display
    rather than a raw URL string spanning a wrapped column. Idempotent on
    a freshly-built workbook: subsequent passes don't re-touch the cell
    because its value is no longer the URL.
    """
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            text = cell.value.strip()
            if text.startswith(("http://", "https://")):
                _set_hyperlink(cell, text)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def _resolve_pinned_datetime(view_graph: rdflib.Graph) -> datetime:
    """Return the deterministic timestamp to use for the xlsx core
    properties — midnight UTC of ``amtmeta:releaseDate`` if present,
    else a fixed sentinel.
    """
    ont = _find_ontology_iri(view_graph)
    release_date_str = _first_value(view_graph, ont, NS_AMTMETA.releaseDate) if ont else ""
    if release_date_str:
        try:
            return datetime.fromisoformat(release_date_str).replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    return datetime(2000, 1, 1, tzinfo=timezone.utc)


def _pin_deterministic_properties(wb: Workbook, view_graph: rdflib.Graph) -> None:
    """Pin :class:`openpyxl.Workbook` core properties to deterministic values.

    Only the *pre-save* fields are settable here — openpyxl re-stamps
    ``modified`` inside its own ``save()`` regardless of any pre-save value,
    so :func:`_overwrite_core_xml_after_save` is the companion fix that
    rewrites ``docProps/core.xml`` after the workbook is on disk.
    """
    pinned_dt = _resolve_pinned_datetime(view_graph)
    wb.properties.creator = "amt-ontology"
    wb.properties.lastModifiedBy = "amt-ontology"
    wb.properties.created = pinned_dt
    wb.properties.modified = pinned_dt


def _overwrite_core_xml_after_save(target: Path, view_graph: rdflib.Graph) -> None:
    """Rewrite ``docProps/core.xml`` inside *target* with a fully pinned
    payload, defeating openpyxl's save-time re-stamping of ``modified``.

    Why this exists: ``wb.save()`` calls
    ``DocumentProperties.modified = datetime.utcnow()`` immediately before
    serializing the core-properties part, overriding any pre-save value we
    set on the property. Pre-save pinning (see
    :func:`_pin_deterministic_properties`) covers ``creator`` / ``created``
    / ``lastModifiedBy``, all of which openpyxl preserves; only ``modified``
    needs this post-save override.

    Without this, three back-to-back ``make dist-xlsx`` runs produce three
    different MD5s and the CI's binary-diff dist-drift check fails on every
    commit. With this, the xlsx bytes are stable as long as the ontology
    inputs (specifically: ``amtmeta:releaseDate``) haven't changed.
    """
    import shutil
    import zipfile

    pinned_iso = _resolve_pinned_datetime(view_graph).strftime("%Y-%m-%dT%H:%M:%SZ")
    core_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        "<cp:coreProperties "
        'xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
        'xmlns:dc="http://purl.org/dc/elements/1.1/" '
        'xmlns:dcterms="http://purl.org/dc/terms/" '
        'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
        "<dc:creator>amt-ontology</dc:creator>"
        f'<dcterms:created xsi:type="dcterms:W3CDTF">{pinned_iso}</dcterms:created>'
        f'<dcterms:modified xsi:type="dcterms:W3CDTF">{pinned_iso}</dcterms:modified>'
        "<cp:lastModifiedBy>amt-ontology</cp:lastModifiedBy>"
        "</cp:coreProperties>\n"
    ).encode()

    # Rebuild the zip with our core.xml replacing the auto-generated one.
    tmp_target = target.with_suffix(target.suffix + ".tmp")
    with (
        zipfile.ZipFile(target, "r") as src,
        zipfile.ZipFile(tmp_target, "w", zipfile.ZIP_DEFLATED) as dst,
    ):
        for item in src.infolist():
            data = core_xml if item.filename == "docProps/core.xml" else src.read(item.filename)
            # Pin ZipInfo timestamps too — without this, the zip's internal
            # file dates drift across runs and add to the binary diff.
            new_info = zipfile.ZipInfo(filename=item.filename, date_time=(1980, 1, 1, 0, 0, 0))
            new_info.compress_type = zipfile.ZIP_DEFLATED
            dst.writestr(new_info, data)
    shutil.move(str(tmp_target), str(target))


def generate_product_interest_xlsx(
    view_graph: rdflib.Graph,
    source_graph: rdflib.Graph,
    amtmeta_graph: rdflib.Graph,
) -> Workbook:
    """Build the Product Interest Category :class:`openpyxl.Workbook`.

    Three graphs are required:

    * *view_graph*    — typically ``visitor_view``; sources the metadata
      header annotations and the bucket hierarchy.
    * *source_graph*  — typically ``pc``; sources the Tier 3
      ``skos:prefLabel`` values for the leaves.
    * *amtmeta_graph* — sources Status individuals' labels (for
      resolving the IRI references in the view header) and the canonical
      Approval Status options for the Look-up values sheet.

    The returned workbook has deterministic ``docProps/core.xml`` so
    that byte-for-byte regeneration is stable as long as the ontology
    inputs haven't changed (see :func:`_pin_deterministic_properties`).
    """
    wb = Workbook()

    md_sheet = wb.active
    md_sheet.title = "Taxonomy Metadata"
    _populate_metadata_sheet(md_sheet, view_graph, amtmeta_graph)

    rl_sheet = wb.create_sheet("Release Log")
    _populate_release_log_sheet(rl_sheet, view_graph, amtmeta_graph)

    hm_sheet = wb.create_sheet("Hierarchy & Mapping")
    _populate_hierarchy_sheet(hm_sheet, view_graph, source_graph)

    lv_sheet = wb.create_sheet("Look-up values")
    _populate_lookup_sheet(lv_sheet, amtmeta_graph)

    # Styling, in order:
    #   1. Hyperlinkify URL cells first — display text becomes the URL's
    #      filename segment, which then drives the auto-sized width.
    #   2. Auto-size every column on every sheet.
    #   3. Wrap the free-form value columns on Taxonomy Metadata (col B)
    #      and Release Log (col C, Scope Summary) so long non-URL values
    #      render cleanly.
    for sheet in (md_sheet, rl_sheet, hm_sheet, lv_sheet):
        _hyperlinkify_url_cells(sheet)
    for sheet in (md_sheet, rl_sheet, hm_sheet, lv_sheet):
        _autosize_columns(sheet)
    _wrap_column(md_sheet, column=2)
    _wrap_column(rl_sheet, column=3)

    _pin_deterministic_properties(wb, view_graph)
    return wb


def workbook_to_bytes(wb: Workbook) -> BytesIO:
    """Serialise *wb* to a ``BytesIO`` buffer positioned at offset 0.

    Mirrors :func:`app.view_csv.rows_to_csv_bytes` so the Flask routes
    layer can serve generated xlsx workbooks the same way it serves
    generated CSV files.
    """
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ---------------------------------------------------------------------------
# CLI entry point — regenerate the dist xlsx
# ---------------------------------------------------------------------------
#
# Invocation::
#
#     python -m app.view_xlsx [output-dir]
#
# When *output-dir* is omitted ``DIST_DIR`` (``./dist/`` by default) is used.
# Used by the ``make dist-xlsx`` target.


def regenerate_dist(out_dir: Path | str | None = None) -> list[Path]:
    """Regenerate the checked-in xlsx export under *out_dir* (default ``dist/``).

    The written file has fully deterministic bytes (see
    :func:`_pin_deterministic_properties` and
    :func:`_overwrite_core_xml_after_save`) so the CI dist-drift check
    passes whenever the ontology hasn't changed.

    Returns the list of files written.
    """
    from .graphs import OntologyGraphs
    from .paths import DIST_DIR

    out_dir = Path(out_dir) if out_dir else DIST_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    graphs = OntologyGraphs()
    wb = generate_product_interest_xlsx(graphs.visitor, graphs.pc, graphs.amtmeta)

    target = out_dir / DIST_FILENAME
    wb.save(target)
    _overwrite_core_xml_after_save(target, graphs.visitor)
    return [target]


def _cli(argv: list[str] | None = None) -> None:
    import argparse
    import sys

    parser = argparse.ArgumentParser(
        prog="python -m app.view_xlsx",
        description="Regenerate the checked-in Product Interest Category xlsx export.",
    )
    parser.add_argument(
        "out_dir",
        nargs="?",
        default=None,
        help="Output directory (defaults to the project's dist/ path).",
    )
    args = parser.parse_args(argv)

    written = regenerate_dist(args.out_dir)
    for p in written:
        print(f"wrote {p}", file=sys.stderr)


if __name__ == "__main__":
    _cli()
