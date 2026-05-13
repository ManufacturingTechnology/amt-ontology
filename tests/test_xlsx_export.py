"""
tests.test_xlsx_export
======================
Functional regression guard for app.view_xlsx — the generator that produces
dist/AMT Taxonomy - Product Interest Category.xlsx.

The generated workbook deliberately diverges from the resource workbook in
ways that don't matter (styling, column widths, multi-row free-form layout,
absence of historical SAMPLE / Previous-versions sheets). What does matter
and what these tests guard is the semantic content.
"""

from __future__ import annotations

import re
import unittest
import warnings
from typing import Any

from openpyxl import load_workbook

from app.graphs import OntologyGraphs
from app.paths import REPO_ROOT
from app.view_xlsx import generate_product_interest_xlsx

RESOURCE_PATH = REPO_ROOT / "resources" / "AMT Taxonomy - Product Interest Category.xlsx"


# ---------------------------------------------------------------------------
# Sheet-reading helpers
# ---------------------------------------------------------------------------


def _read_sheet_rows(ws) -> list[tuple[Any, ...]]:
    rows: list[tuple[Any, ...]] = []
    for row in ws.iter_rows(values_only=True):
        trimmed = list(row)
        while trimmed and (trimmed[-1] is None or trimmed[-1] == ""):
            trimmed.pop()
        rows.append(tuple(trimmed))
    return rows


def _flatten_metadata(rows):
    grouped: dict[str, list[str]] = {}
    current_label: str | None = None
    for row in rows:
        if not row:
            continue
        label = row[0]
        value = row[1] if len(row) > 1 else ""
        if label:
            current_label = str(label)
            grouped.setdefault(current_label, [])
            if value not in (None, ""):
                grouped[current_label].append(str(value).strip())
        else:
            if current_label and value not in (None, ""):
                grouped[current_label].append(str(value).strip())
    return grouped


def _extract_hierarchy_rows(rows):
    """Return set of (Tier1, Tier2, Tier3) tuples, header rows skipped."""
    result: set[tuple[str, str, str]] = set()
    for row in rows:
        if not row:
            continue
        first = row[0]
        if first in ("Current version", "Entry ID"):
            continue
        padded = list(row) + [None] * 5

        def _s(x):
            if x is None:
                return ""
            return str(x).strip()

        tier1, tier2, tier3 = _s(padded[1]), _s(padded[2]), _s(padded[3])
        if not (tier1 or tier2 or tier3):
            continue
        result.add((tier1, tier2, tier3))
    return result


def _normalise_use_case(text):
    return text.strip().rstrip(".").strip().lower()


def _word_set(text):
    return {tok for tok in re.findall(r"\w+", text.lower()) if tok}


def _jaccard(a, b):
    wa, wb = _word_set(a), _word_set(b)
    if not wa or not wb:
        return 0.0
    return len(wa & wb) / len(wa | wb)


def _fuzzy_in(needle, haystack, *, threshold=0.7):
    if needle in haystack:
        return True
    return any(_jaccard(needle, h) >= threshold for h in haystack)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


class TestProductInterestCategoryXlsx(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        if not RESOURCE_PATH.exists():
            raise unittest.SkipTest(f"Baseline xlsx missing: {RESOURCE_PATH}")
        graphs = OntologyGraphs()
        cls.generated_wb = generate_product_interest_xlsx(graphs.visitor, graphs.pc, graphs.amtmeta)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            cls.resource_wb = load_workbook(RESOURCE_PATH, data_only=True)

    def test_required_sheets_present(self):
        required = {"Taxonomy Metadata", "Release Log", "Hierarchy & Mapping", "Look-up values"}
        gen = set(self.generated_wb.sheetnames)
        missing = required - gen
        self.assertFalse(missing, f"Missing sheets: {sorted(missing)}")

    def test_taxonomy_metadata_required_labels(self):
        gen_md = _flatten_metadata(_read_sheet_rows(self.generated_wb["Taxonomy Metadata"]))
        required = {
            "Taxonomy name:",
            "Taxonomy owner:",
            "Status:",
            "Governers:",
            "Consumers:",
            "Tech stewards:",
            "Related products:",
            "Related systems/databases:",
            "Next update date:",
            "Related documentation:",
            "Use cases:",
            "Notes:",
        }
        missing = required - set(gen_md.keys())
        self.assertFalse(missing, f"Missing labels: {sorted(missing)}")

    def test_taxonomy_metadata_status_matches_resource(self):
        gen = _flatten_metadata(_read_sheet_rows(self.generated_wb["Taxonomy Metadata"]))
        res = _flatten_metadata(_read_sheet_rows(self.resource_wb["Taxonomy Metadata"]))
        g = " ".join(gen.get("Status:", [])).strip().lower()
        r = " ".join(res.get("Status:", [])).strip().lower()
        self.assertEqual(g, r, f"Status mismatch: gen={g!r} res={r!r}")

    def test_taxonomy_metadata_owner_matches_resource(self):
        gen = _flatten_metadata(_read_sheet_rows(self.generated_wb["Taxonomy Metadata"]))
        res = _flatten_metadata(_read_sheet_rows(self.resource_wb["Taxonomy Metadata"]))
        g = " ".join(gen.get("Taxonomy owner:", [])).strip().lower()
        r = " ".join(res.get("Taxonomy owner:", [])).strip().lower()
        self.assertEqual(g, r, f"Owner mismatch: gen={g!r} res={r!r}")

    def test_taxonomy_metadata_use_cases_overlap_with_resource(self):
        """Resource use cases must be findable in the generated set with
        Jaccard ≥ 0.7 (tolerates typos like 'Passprt' vs 'Passport')."""
        gen = _flatten_metadata(_read_sheet_rows(self.generated_wb["Taxonomy Metadata"]))
        res = _flatten_metadata(_read_sheet_rows(self.resource_wb["Taxonomy Metadata"]))
        gen_uc = {_normalise_use_case(s) for s in gen.get("Use cases:", []) if s}
        res_uc = {_normalise_use_case(s) for s in res.get("Use cases:", []) if s}
        uncovered = {uc for uc in res_uc if not _fuzzy_in(uc, gen_uc) and "future" not in uc}
        self.assertFalse(
            uncovered,
            f"Resource use cases not represented ({len(uncovered)}): "
            + "; ".join(sorted(uncovered)),
        )

    def test_lookup_values_exactly_match_resource(self):
        def value_set(rows):
            return {str(c).strip() for row in rows[1:] for c in row if c not in (None, "")}

        gen_set = value_set(_read_sheet_rows(self.generated_wb["Look-up values"]))
        res_set = value_set(_read_sheet_rows(self.resource_wb["Look-up values"]))
        self.assertEqual(
            gen_set, res_set, f"Lookup values: gen={sorted(gen_set)} res={sorted(res_set)}"
        )

    def test_release_log_headers_match_resource(self):
        gen = _read_sheet_rows(self.generated_wb["Release Log"])
        res = _read_sheet_rows(self.resource_wb["Release Log"])
        self.assertEqual(
            list(gen[0]),
            list(res[0]),
            f"Release Log headers: gen={list(gen[0])} res={list(res[0])}",
        )

    def test_hierarchy_mapping_tier3_leaves_match_resource(self):
        """Tier 3 (the leaf Product Category) is the semantically load-bearing
        column — its set must match the resource exactly."""
        gen_rows = _extract_hierarchy_rows(
            _read_sheet_rows(self.generated_wb["Hierarchy & Mapping"])
        )
        res_rows = _extract_hierarchy_rows(
            _read_sheet_rows(self.resource_wb["Hierarchy & Mapping"])
        )
        gen_t3 = {r[2] for r in gen_rows if r[2]}
        res_t3 = {r[2] for r in res_rows if r[2]}
        only_gen = gen_t3 - res_t3
        only_res = res_t3 - gen_t3
        if not (only_gen or only_res):
            return
        msg = [f"Tier 3 leaf sets differ; gen={len(gen_t3)} res={len(res_t3)}"]
        if only_gen:
            msg.append(f"  in generated only ({len(only_gen)}): {sorted(only_gen)[:10]}")
        if only_res:
            msg.append(f"  in resource only ({len(only_res)}): {sorted(only_res)[:10]}")
        self.fail("\n".join(msg))

    def test_hierarchy_mapping_grouping_structure_matches_resource(self):
        """Leaf-grouping must be preserved even when Tier 1 labels change
        (e.g. 'Industrial Artificial Intelligence' -> 'Industrial AI'):
        the SAME four leaves must still appear under SOME Tier 1 grouping."""
        gen_rows = _extract_hierarchy_rows(
            _read_sheet_rows(self.generated_wb["Hierarchy & Mapping"])
        )
        res_rows = _extract_hierarchy_rows(
            _read_sheet_rows(self.resource_wb["Hierarchy & Mapping"])
        )

        def groups_by_tier1(rows):
            out: dict[str, set[str]] = {}
            for t1, _t2, t3 in rows:
                if not t1 or not t3:
                    continue
                out.setdefault(t1, set()).add(t3)
            return {k: frozenset(v) for k, v in out.items()}

        gen_groups = set(groups_by_tier1(gen_rows).values())
        res_groups = groups_by_tier1(res_rows)
        unmatched = {leaves for leaves in res_groups.values() if leaves not in gen_groups}
        if not unmatched:
            return
        leaves_to_label = {leaves: t1 for t1, leaves in res_groups.items()}
        msg = [f"{len(unmatched)} resource Tier 1 grouping(s) have no match in generated"]
        for leaves in list(unmatched)[:5]:
            label = leaves_to_label.get(leaves, "?")
            msg.append(f"  - resource '{label}': {sorted(leaves)[:5]}")
        self.fail("\n".join(msg))


if __name__ == "__main__":
    unittest.main(verbosity=2)
